import re
import gc
from pathlib import Path
from datetime import datetime, date
from src.logger_manager import get_logger

log = get_logger("FileExtractor")

class FileExtractor:
    def __init__(self, root_directory):
        self.root_directory = Path(root_directory)

    def extract_chunks_for_month(self, target_month, target_year):
        from openpyxl import load_workbook
        
        date_pattern = re.compile(r"^\d{2}-\d{2}-\d{4}$")
        log.info(f"Executing ultra-low memory directory scan for timeframe: {target_year}-{target_month:02d}")
        
        valid_folder_discovered = False

        for folder in self.root_directory.iterdir():
            if not folder.is_dir() or not date_pattern.match(folder.name):
                continue

            try:
                folder_date = datetime.strptime(folder.name, "%d-%m-%Y")
            except ValueError:
                log.warning(f"Skipping incorrectly formatted date folder structure: {folder.name}")
                continue

            if folder_date.month != target_month or folder_date.year != target_year:
                continue

            valid_folder_discovered = True
            log.info(f"Target chronological directory matched: {folder.name}")
            
            excel_files = [f for f in folder.iterdir() if f.is_file() and f.suffix in ['.xlsx', '.xls']]
            
            if not excel_files:
                log.warning(f"No executable spreadsheet files discovered inside directory: {folder.name}")
                continue
                
            if len(excel_files) > 1:
                raise ValueError(f"Ambiguity conflict in directory [{folder.name}]: Multiple data payloads detected.")

            target_file_path = excel_files[0]
            log.info(f"Streaming data stream from binary spreadsheet: {target_file_path.name}")
            
            try:
                workbook = load_workbook(
                    filename=str(target_file_path), 
                    read_only=True, 
                    data_only=True,
                    keep_links=False,
                    keep_vba=False
                )
                worksheet = workbook.active
                
                yield from self._parse_structured_layout(worksheet)
                
                workbook.close()
                del worksheet
                del workbook
                gc.collect()
            except Exception as e:
                log.error(f"Critical operational failure deserializing data file: {target_file_path.name}")
                raise e

        if not valid_folder_discovered:
            raise FileNotFoundError(f"No matching operational target directories found for timeframe: {target_year}-{target_month:02d}")

    def _parse_structured_layout(self, worksheet):
        current_employee = None
        current_id = None
        target_col_idx = None
        dash_date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}.*")
        
        for row in worksheet.iter_rows(values_only=True):
            if not row:
                continue
                
            row_values = []
            for val in row:
                if val is None:
                    row_values.append("")
                elif isinstance(val, (datetime, date)):
                    row_values.append(val.strftime("%m/%d/%Y"))
                else:
                    row_values.append(str(val).strip())
            
            employee_cell_idx = None
            for idx, val in enumerate(row_values):
                if val.startswith("Employee:"):
                    employee_cell_idx = idx
                    break
            
            if employee_cell_idx is not None:
                if current_employee:
                    yield {
                        "EMPLOYEE": current_employee,
                        "EMPLOYEE ID": current_id,
                        "START DATE": ""
                    }
                
                current_employee = row_values[employee_cell_idx].replace("Employee:", "").strip()
                current_id = ""
                target_col_idx = None
                
                for idx, val in enumerate(row_values):
                    if "Employee ID:" in val:
                        current_id = val.split("Employee ID:")[-1].strip()
                    if "Supervisor:" in val:
                        target_col_idx = idx
                continue

            if current_employee and target_col_idx is not None and target_col_idx < len(row_values):
                raw_date_cell = row_values[target_col_idx]
                start_date_value = ""
                
                if "/" in raw_date_cell:
                    start_date_value = raw_date_cell.split(" ")[0].strip()
                elif dash_date_pattern.match(raw_date_cell):
                    try:
                        clean_date = raw_date_cell.split(" ")[0].strip()
                        dt = datetime.strptime(clean_date, "%Y-%m-%d")
                        start_date_value = dt.strftime("%m/%d/%Y")
                    except ValueError:
                        pass
                
                if start_date_value:
                    yield {
                        "EMPLOYEE": current_employee,
                        "EMPLOYEE ID": current_id,
                        "START DATE": start_date_value
                    }
                    current_employee = None
                    current_id = None
                    target_col_idx = None

        if current_employee:
            yield {
                "EMPLOYEE": current_employee,
                "EMPLOYEE ID": current_id,
                "START DATE": ""
            }


